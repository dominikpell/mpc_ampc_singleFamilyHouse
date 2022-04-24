import pandas as pd
import DataTuning
import ModelTuning
import os
import SharedVariables as SV
from openpyxl import load_workbook

# Define path to excel file containing experimental plan
PathToExperimentalPlan = r'D:\05_Offline\AI_Paper\addmo-automated-ml-regression\ExperimentalPlan.xlsx'

# Define name of excel sheet
sheet = '18MonthsTraining_6MonthsTesting'

### Read information from experimental plan ###
# Read general data of experimental plan
#dfGeneralData = pd.read_excel(PathToExperimentalPlan, sheet_name=sheet, usecols=[10, 11],
 #                             skiprows=lambda x: x not in [2, 3, 4, 5, 6, 7], names=['event', 'date'], index_col='event', )
# Load excel file containing experimental plan as workbook
wb = load_workbook(PathToExperimentalPlan)
ws = wb[sheet]

# save general data correctly formatted in "SharedVariables"
#SV.StartTraining = '2018-08-01 00:00'
#SV.EndTraining = "2018-08-12 00:00"
#SV.StartTesting = "2018-08-12 00:00"
#SV.EndTesting = "2018-08-16 23:30:00"

# read parameters of experimental plan
dfParameters = pd.read_excel(PathToExperimentalPlan, sheet_name=[sheet], header=1, index_col='Nr.',
                             usecols=['Nr.', 'GlobalMaxEval_HyParaTuning', 'MaxEval_Bayes', 'TrainingStart', 'TrainingEnd', 'TestStart', 'TestEnd'])
dfParameters = dfParameters[sheet]
#print(dfParameters)

# Set score type
SV.score_type = 'R^2 '  # or 'RMSE'
#Whether data tuning needs to be performed or not
data_tune_true = False
k = SV.score_type

#Iterate over experimental plan excel sheet
for i in range(1, 1 + len(dfParameters.index), 1):  # for each combination of parameters

    #Model tuning relevant parameters
    SV.GlobalMaxEval_HyParaTuning = dfParameters['GlobalMaxEval_HyParaTuning'][i]
    SV.MaxEval_Bayes = dfParameters['MaxEval_Bayes'][i]

    # save general data correctly formatted in "SharedVariables"
    SV.StartTraining = str(dfParameters['TrainingStart'][i])
    SV.EndTraining = str(dfParameters['TrainingEnd'][i])
    SV.StartTesting = str(dfParameters['TestStart'][i])
    SV.EndTesting = str(dfParameters['TestEnd'][i])

    ### Set names of folders and paths ###
    # Set unique name of experiment
    SV.NameOfExperiment = "nFtLag_nIndMod_nOL_QSet_ANNHL16N_b"
    SV.NameOfData = "AI Paper"
    print(SV.NameOfData)
    print(SV.NameOfExperiment)

    if data_tune_true == True:
        DataTuning.main()  # this function includes: Set general folder for results, which are used in each experiment
    else:
        None

    # Define path to data source files '.xls' & '.pickle'
    RootDir = os.path.dirname(os.path.realpath(__file__))
    PathToData = os.path.join(RootDir, 'Data')

    # Set specific folder for results of experiment
    ResultsFolder = os.path.join(RootDir, "Results", SV.NameOfData, SV.NameOfExperiment, str(i))
    PathToPickles = os.path.join(ResultsFolder, "Pickles")
    print(ResultsFolder)

    # Set folders and paths in "SharedVariables"
    SV.RootDir = RootDir
    SV.PathToData = PathToData
    SV.ResultsFolder = ResultsFolder
    SV.PathToPickles = PathToPickles
try:
    # Execute ModelTuning (dependent of chosen parameters)
    # main_OnlyHyParaOpti (only model tuning or main_FinalBayes ( if both Data and Model Tuning should be performed)
    ModelTuning.main_OnlyHyParaOpti()

    kpi = SV.output_value
    print(kpi)
    # Save output in workbook
    if k == 'R^2 ':
        column = 9
    elif k == 'RMSE':
        column = 10
    else:
        raise NameError(
            'The selected score_type needs to be assigned to a column in the excel file containing the experimental plan')
    ws.cell(row=i + 2, column=column, value=SV.output_value)

# Save workbook with outputs in excel file containing experimental plan
wb.save(PathToExperimentalPlan)
wb.close()