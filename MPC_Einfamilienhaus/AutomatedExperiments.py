import pandas as pd
import DataTuning
import ModelTuning
import os
import SharedVariables as SV
from openpyxl import load_workbook
import random


### User input ###
# Define path to excel file containing experimental plan
PathToExperimentalPlan = r'D:\05_Offline\AI_Paper\addmo-automated-ml-regression\ExperimentalPlan.xlsx'

# Define name of excel sheet
sheet = '18MonthsTraining_6MonthsTesting'

# Define score types, which are used for tuning and to evaluate each experiment
all_score_types = ['R^2 ', 'RMSE']  # in order to add other score types the corresponding functions in "ModelTuning" need to be adapted


### Read information from experimental plan ###
# Read general data of experimental plan
dfGeneralData = pd.read_excel(PathToExperimentalPlan, sheet_name=sheet, usecols=[10, 11],
                              skiprows=lambda x: x not in [2, 3, 4, 5, 6], names=['event', 'date'], index_col='event', )
print(dfGeneralData)

# save general data correctly formatted in "SharedVariables"
SV.StartTraining = str(dfGeneralData['date']['Training Start:']).replace('00:00:00', '00:00')
SV.EndTraining = str(dfGeneralData['date']['Training End:']).replace('00:00:00', '23:45')
SV.StartTesting = str(dfGeneralData['date']['Test Start:']).replace('00:00:00', '00:00')
SV.EndTesting = str(dfGeneralData['date']['Test End:']).replace('00:00:00', '23:45')

# read parameters of experimental plan
#dfParameters = pd.read_excel(PathToExperimentalPlan, sheet_name=[sheet], header=1, index_col='Nr.',
                             #usecols=['Nr.', 'n_estimators', 'max_depth', 'max_leaf_nodes', 'random_state', 'GlobalMaxEval_HyParaTuning'])
#dfParameters = dfParameters[sheet]
#print(dfParameters)


# Load excel file containing experimental plan as workbook
wb = load_workbook(PathToExperimentalPlan)
ws = wb[sheet]

# Execute DataTuning once for all experiments (independent of chosen parameters)
DataTuning.main()  # this function includes: Set general folder for results, which are used in each experiment


# TESTING ONLY
def test_automation(n_estimators, max_depth, max_leaf_nodes, random_state, GlobalMaxEval_HyParaTuning, score_type, number):
    """can be used for testing only to replace real experiment execution to save computing time"""
    print("Skript des", number,
          "Versuchs mit n_estimators =", n_estimators,
          "und mit max_depth =", max_depth,
          "und mit max_leaf_nodes =", max_leaf_nodes,
          "und mit random_state =", random_state,
          "und mit GlobalMaxEval_HyParaTuning =", GlobalMaxEval_HyParaTuning,
          "für den Score", score_type)


# Execute all experiments automatically
for i in range(1, 1 + len(dfParameters.index), 1):  # for each combination of parameters
    for k in all_score_types:  # for each score type
        # Set score type
        SV.score_type = k

        # Set parameters in "SharedVariables" (overwriting last value)
        SV.n_estimators = dfParameters['n_estimators'][i]
        SV.max_depth = dfParameters['max_depth'][i]
        SV.max_leaf_nodes = dfParameters['max_leaf_nodes'][i]
        SV.random_state = dfParameters['random_state'][i]
        SV.GlobalMaxEval_HyParaTuning = dfParameters['GlobalMaxEval_HyParaTuning'][i]


        ### Set names of folders and paths ###
        # Set unique name of experiment
        SV.NameOfExperiment = 'Nr%s_score%s_nest%s_maxde%s_maxleaves%s_ransta%s_maxEval%s'\
                              % (i, k, SV.n_estimators, SV.max_depth, SV.max_leaf_nodes, SV.random_state, SV.GlobalMaxEval_HyParaTuning)
        print(SV.NameOfExperiment)

        # Define path to data source files '.xls' & '.pickle'
        RootDir = os.path.dirname(os.path.realpath(__file__))
        PathToData = os.path.join(RootDir, 'Data')

        # Set specific folder for results of experiment
        ResultsFolder = os.path.join(RootDir, "Results", SV.NameOfData, SV.NameOfExperiment)
        PathToPickles = os.path.join(ResultsFolder, "Pickles")

        # Set folders and paths in "SharedVariables"
        SV.RootDir = RootDir
        SV.PathToData = PathToData
        SV.ResultsFolder = ResultsFolder
        SV.PathToPickles = PathToPickles


        # Execute ModelTuning (dependent of chosen parameters)
        ModelTuning.main_FinalBayes()

        # Save output in workbook
        if k == 'R^2 ':
            column = 8
        elif k == 'RMSE':
            column = 9
        else:
            raise NameError('The selected score_type needs to be assigned to a column in the excel file containing the experimental plan')
        ws.cell(row=i + 2, column=column, value=SV.output_value)


# Save workbook with outputs in excel file containing experimental plan
wb.save(PathToExperimentalPlan)
wb.close()
